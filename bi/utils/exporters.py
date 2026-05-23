"""
Module d'Export Avancé
CSV, Excel, PDF pour E-GovProjetGB
"""

import pandas as pd
import numpy as np
from datetime import datetime
from typing import Dict, Any, List, Optional
from io import BytesIO
import logging

logger = logging.getLogger(__name__)

# ============================================================================
# EXPORT CSV
# ============================================================================

class CSVExporter:
    """Exporte les données en format CSV optimisé"""
    
    @staticmethod
    def export_projects(
        projects_df: pd.DataFrame,
        include_columns: Optional[List[str]] = None,
        format_currency: bool = True,
        encoding: str = 'utf-8-sig'  # UTF-8 with BOM for Excel
    ) -> str:
        """
        Exporter projets en CSV
        
        Args:
            projects_df: DataFrame des projets
            include_columns: Colonnes à inclure (None = toutes)
            format_currency: Formater les montants (XOF)
            encoding: Encoding du fichier
        
        Returns:
            CSV string
        """
        
        df = projects_df.copy()
        
        # Sélectionner colonnes
        if include_columns:
            cols = [c for c in include_columns if c in df.columns]
            df = df[cols]
        
        # Formater dates
        date_cols = df.select_dtypes(include=['datetime64']).columns
        for col in date_cols:
            df[col] = df[col].dt.strftime('%d/%m/%Y')
        
        # Formater monnaie
        if format_currency:
            currency_cols = [c for c in df.columns if 'xof' in c.lower()]
            for col in currency_cols:
                df[col] = df[col].apply(lambda x: f"{x:,.0f}" if pd.notna(x) else "")
        
        # Formater pourcentages
        pct_cols = [c for c in df.columns if 'progress' in c.lower() or 'variance' in c.lower()]
        for col in pct_cols:
            if col in df.columns:
                df[col] = df[col].apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "")
        
        # Exporter CSV
        csv = df.to_csv(index=False, encoding=encoding)
        logger.info(f"CSV exported: {len(df)} projects, {len(df.columns)} columns")
        
        return csv
    
    @staticmethod
    def export_kpis(
        kpis_dict: Dict[str, Any],
        metadata: Optional[Dict[str, Any]] = None
    ) -> str:
        """
        Exporter KPIs en CSV
        
        Args:
            kpis_dict: Dictionnaire des KPIs
            metadata: Métadonnées (date, région, etc.)
        
        Returns:
            CSV string
        """
        
        rows = []
        
        # Header
        if metadata:
            for key, value in metadata.items():
                rows.append([key, value])
            rows.append([])  # Ligne vide
        
        # KPIs
        for kpi_name, kpi_data in kpis_dict.items():
            if isinstance(kpi_data, dict):
                rows.append([kpi_name])
                for key, value in kpi_data.items():
                    if not key.startswith('_'):
                        rows.append(['  ' + key, value])
            else:
                rows.append([kpi_name, kpi_data])
        
        df = pd.DataFrame(rows, columns=['Indicateur', 'Valeur'])
        csv = df.to_csv(index=False, encoding='utf-8-sig')
        
        return csv
    
    @staticmethod
    def export_audit_logs(
        audit_logs: pd.DataFrame,
        include_columns: Optional[List[str]] = None
    ) -> str:
        """
        Exporter journaux d'audit en CSV
        
        Args:
            audit_logs: DataFrame des logs d'audit
            include_columns: Colonnes à inclure
        
        Returns:
            CSV string
        """
        
        df = audit_logs.copy()
        
        if include_columns:
            cols = [c for c in include_columns if c in df.columns]
            df = df[cols]
        
        # Formater timestamps
        if 'created_at' in df.columns:
            df['created_at'] = df['created_at'].dt.strftime('%d/%m/%Y %H:%M:%S')
        
        csv = df.to_csv(index=False, encoding='utf-8-sig')
        logger.info(f"Audit logs exported: {len(df)} entries")
        
        return csv

# ============================================================================
# EXPORT EXCEL
# ============================================================================

class ExcelExporter:
    """Exporte les données en format Excel avec formatage"""
    
    @staticmethod
    def export_projects(
        projects_df: pd.DataFrame,
        include_columns: Optional[List[str]] = None
    ) -> bytes:
        """
        Exporter projets en Excel avec formatage
        
        Args:
            projects_df: DataFrame des projets
            include_columns: Colonnes à inclure
        
        Returns:
            Bytes (fichier Excel)
        """
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            logger.error("openpyxl not installed. Install with: pip install openpyxl")
            raise
        
        df = projects_df.copy()
        
        # Sélectionner colonnes
        if include_columns:
            cols = [c for c in include_columns if c in df.columns]
            df = df[cols]
        
        # Créer workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Projets"
        
        # Écrire données
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Formatage header
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Formatage données
                if r_idx > 1:
                    # Monnaie
                    if 'xof' in str(cell.column_letter).lower() or any(x in str(value).lower() for x in ['xof', 'budget', 'spent']):
                        cell.number_format = '#,##0'
                    
                    # Pourcentage
                    if any(x in str(cell.coordinate).lower() for x in ['progress', 'variance']):
                        cell.number_format = '0.0"%"'
                    
                    # Date
                    if isinstance(value, (pd.Timestamp, datetime)):
                        cell.number_format = 'DD/MM/YYYY'
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze panes
        ws.freeze_panes = "A2"
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"Excel exported: {len(df)} projects")
        return output.getvalue()
    
    @staticmethod
    def export_report(
        report_data: Dict[str, Any],
        title: str = "Rapport E-GovProjetGB"
    ) -> bytes:
        """
        Exporter rapport complet en Excel multi-feuilles
        
        Args:
            report_data: Dict avec clés = noms feuilles, valeurs = DataFrames
            title: Titre du rapport
        
        Returns:
            Bytes (fichier Excel)
        """
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            logger.error("openpyxl not installed")
            raise
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Créer une feuille pour chaque section
        for sheet_name, data in report_data.items():
            ws = wb.create_sheet(sheet_name)
            
            if isinstance(data, pd.DataFrame):
                # Écrire DataFrame
                from openpyxl.utils.dataframe import dataframe_to_rows
                
                for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        
                        if r_idx == 1:
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            elif isinstance(data, dict):
                # Écrire dictionnaire (KPIs, stats)
                for r_idx, (key, value) in enumerate(data.items(), 1):
                    ws.cell(row=r_idx, column=1, value=key).font = Font(bold=True)
                    ws.cell(row=r_idx, column=2, value=value)
        
        # Save
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"Report exported with {len(report_data)} sheets")
        return output.getvalue()

# ============================================================================
# EXPORT PDF
# ============================================================================

class PDFExporter:
    """Exporte les données en format PDF"""
    
    @staticmethod
    def export_summary_report(
        projects_df: pd.DataFrame,
        kpis: Dict[str, Any],
        title: str = "Rapport Synthétique E-GovProjetGB",
        subtitle: str = "",
        include_tables: bool = True
    ) -> bytes:
        """
        Exporter rapport synthétique en PDF
        
        Args:
            projects_df: DataFrame des projets
            kpis: Dictionnaire des KPIs
            title: Titre du rapport
            subtitle: Sous-titre
            include_tables: Inclure tableaux détaillés
        
        Returns:
            Bytes (fichier PDF)
        """
        
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        except ImportError:
            logger.error("reportlab not installed. Install with: pip install reportlab")
            raise
        
        # Create PDF
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=0.5*inch, leftMargin=0.5*inch,
                               topMargin=0.75*inch, bottomMargin=0.75*inch)
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=6,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        elements.append(Paragraph(title, title_style))
        
        # Subtitle
        if subtitle:
            subtitle_style = ParagraphStyle(
                'Subtitle',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.HexColor('#6b7280'),
                spaceAfter=12,
                alignment=TA_CENTER
            )
            elements.append(Paragraph(subtitle, subtitle_style))
        
        # Metadata
        elements.append(Paragraph(f"<b>Date:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))
        
        # KPIs Section
        elements.append(Paragraph("<b>📊 Indicateurs Clés de Performance</b>", styles['Heading2']))
        elements.append(Spacer(1, 0.1*inch))
        
        kpi_data = [['Indicateur', 'Valeur', 'Statut']]
        for kpi_name, kpi_value in kpis.items():
            if isinstance(kpi_value, dict):
                value = kpi_value.get('value', kpi_value.get('variance', ''))
                status = kpi_value.get('status', '')
                kpi_data.append([kpi_name, str(value), str(status)])
            else:
                kpi_data.append([kpi_name, str(kpi_value), ''])
        
        kpi_table = Table(kpi_data, colWidths=[3*inch, 1.5*inch, 1*inch])
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        elements.append(kpi_table)
        elements.append(Spacer(1, 0.2*inch))
        
        # Projects Summary
        elements.append(Paragraph("<b>📋 Résumé Projets</b>", styles['Heading2']))
        elements.append(Spacer(1, 0.1*inch))
        
        summary_data = [
            ['Total Projets', str(len(projects_df))],
            ['Actifs', str(len(projects_df[projects_df['status'].isin(['PLANNED', 'IN_PROGRESS', 'BLOCKED'])]))],
            ['Complétés', str(len(projects_df[projects_df['status'] == 'COMPLETED']))],
            ['Budget Total', f"{projects_df['budget_xof'].sum():,.0f} XOF"],
            ['Dépensé', f"{projects_df['spent_xof'].sum():,.0f} XOF"],
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2.5*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e5e7eb')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 0.2*inch))
        
        # Projects Table
        if include_tables and not projects_df.empty:
            elements.append(PageBreak())
            elements.append(Paragraph("<b>📊 Détail Projets</b>", styles['Heading2']))
            elements.append(Spacer(1, 0.1*inch))
            
            # Top 10 projects by budget
            top_projects = projects_df.nlargest(10, 'budget_xof')[
                ['name', 'region', 'sector', 'budget_xof', 'spent_xof', 'progress', 'status']
            ].copy()
            
            table_data = [['Projet', 'Région', 'Secteur', 'Budget (XOF)', 'Dépensé (XOF)', 'Avancement', 'Statut']]
            for _, row in top_projects.iterrows():
                table_data.append([
                    row['name'][:30],
                    row['region'],
                    row['sector'],
                    f"{row['budget_xof']:,.0f}",
                    f"{row['spent_xof']:,.0f}",
                    f"{row['progress']:.0f}%",
                    row['status']
                ])
            
            projects_table = Table(table_data, colWidths=[1.5*inch, 1*inch, 1*inch, 0.9*inch, 0.9*inch, 0.8*inch, 0.8*inch])
            projects_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            elements.append(projects_table)
        
        # Footer
        elements.append(Spacer(1, 0.3*inch))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#9ca3af'),
            alignment=TA_CENTER
        )
        elements.append(Paragraph("E-GovProjetGB • Plateforme de Gouvernance Publique • Guinée-Bissau 🇬🇼", footer_style))
        
        # Build PDF
        doc.build(elements)
        output.seek(0)
        
        logger.info("PDF report generated successfully")
        return output.getvalue()

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def get_export_filename(export_type: str, suffix: str = "") -> str:
    """Générer nom de fichier pour export"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = f"egov_projets_{suffix}_{timestamp}" if suffix else f"egov_projets_{timestamp}"
    
    extensions = {
        'csv': 'csv',
        'excel': 'xlsx',
        'pdf': 'pdf'
    }
    
    ext = extensions.get(export_type, 'txt')
    return f"{base}.{ext}"
